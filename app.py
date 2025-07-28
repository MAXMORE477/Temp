from flask import Flask, request, jsonify, abort
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from urllib.parse import quote

# Load environment variables from .env
load_dotenv()
API_KEY = os.getenv("API_KEY")

# Config
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
PER_PAGE = 1000

app = Flask(__name__)
limiter = Limiter(get_remote_address, app=app, default_limits=["100 per hour"])

# 🔐 Authentication middleware
def require_api_key():
    token = request.headers.get("Authorization")
    if token != f"Bearer {API_KEY}":
        abort(401, description="Unauthorized")

# 📁 List all available Excel files
@app.route("/files")
def list_files():
    require_api_key()
    files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    return jsonify({"files": files})

# 📄 Get paginated data from the only sheet in the file
@app.route("/file/<filename>")
def get_sheet_data(filename):
    require_api_key()
    filepath = os.path.join(DATA_DIR, filename)
    if not os.path.isfile(filepath):
        return jsonify({"error": "File not found"}), 404

    try:
        page = int(request.args.get("page", 1))
        if page <= 0:
            raise ValueError
    except ValueError:
        return jsonify({"error": "Invalid page value"}), 400

    try:
        wb = load_workbook(filepath, read_only=True)
        sheet = wb.worksheets[0]  # Assume only one sheet per file

        # Read header
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        total_rows = sheet.max_row - 1  # excluding header
        total_pages = (total_rows + PER_PAGE - 1) // PER_PAGE

        start_row = (page - 1) * PER_PAGE + 2  # skip header row
        end_row = start_row + PER_PAGE - 1

        data = []
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
            values = [cell.value for cell in row]
            if any(v is not None for v in values):  # avoid empty rows
                data.append(dict(zip(header, values)))

        has_more = end_row < sheet.max_row
        next_page = page + 1 if has_more else None

        return jsonify({
            "file": filename,
            "page": page,
            "per_page": PER_PAGE,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "has_more": has_more,
            "next_page": next_page,
            "data": data
        })

    except Exception as e:
        return jsonify({"error": f"Could not load data: {str(e)}"}), 500

# 🔐 Error handlers
@app.errorhandler(401)
def unauthorized(e):
    return jsonify({"error": str(e)}), 401

@app.errorhandler(429)
def rate_limit_exceeded(e):
    return jsonify({"error": "Rate limit exceeded"}), 429

if __name__ == "__main__":
    app.run(debug=True)
